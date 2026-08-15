"""
REST API for Season restaurant reservation system.
"""
import csv
import io
import threading
from functools import wraps
from flask import Blueprint, request, jsonify, current_app, Response, session
from datetime import date, timedelta, datetime
from services import reservation as res_svc
from services import client as cli_svc
from services import reports as rpt_svc
from models import db, Reservation, Waitlist, Table
from utils.validators import ReservationCreate, ReservationUpdate, ClientCreate, ClientUpdate
from utils.decorators import validate_and_handle, handle_errors
from utils.errors import error_response


def _send_email_async(app, reservation_data):
    """Fire-and-forget: send confirmation email in a background thread."""
    def _run():
        with app.app_context():
            try:
                from services.email_service import send_confirmation_email
                send_confirmation_email(
                    to_email=reservation_data.get('email', ''),
                    name=reservation_data.get('client_name', ''),
                    date_str=reservation_data.get('date', ''),
                    shift=reservation_data.get('shift', ''),
                    time_str=reservation_data.get('time', ''),
                    guests=reservation_data.get('guests', 1),
                    notes=reservation_data.get('notes', ''),
                )
            except Exception:
                pass
    threading.Thread(target=_run, daemon=True).start()

api = Blueprint('api', __name__, url_prefix='/api')


@api.before_request
def require_login():
    """Require authenticated session for all API endpoints."""
    # Allow health check and public booking unauthenticated
    if request.endpoint in ('api.health', 'api.public_reserve'):
        return None
    if not session.get('user_id'):
        return jsonify({'error': 'No autenticado', 'code': 'UNAUTHENTICATED'}), 401


def notify(date_str=None, shift=None):
    from app import broadcast_update, cache_invalidate
    cache_invalidate(date_str, shift)
    broadcast_update()


def cached_report(max_age=60):
    """Cache a GET report endpoint by its full path+querystring.

    Report data changes only when reservations change, and notify()/
    cache_invalidate() already purges every key starting with 'reports'.
    So a short TTL is safe and cuts repeated aggregation queries against the
    remote Postgres when the Informes view switches modes (hoy/semana/mes)
    or several devices open reports at once.
    """
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            from app import cache_get, cache_set
            key = f'reports:{request.full_path}'
            cached = cache_get(key, max_age=max_age)
            if cached is not None:
                return current_app.response_class(cached, mimetype='application/json')
            resp = fn(*args, **kwargs)
            try:
                r = resp[0] if isinstance(resp, tuple) else resp
                if getattr(r, 'status_code', 200) == 200:
                    cache_set(key, r.get_data(as_text=True))
            except Exception:
                pass
            return resp
        return wrapper
    return decorator


# ── Reservations ──────────────────────────────────────────

@api.route('/reservations', methods=['GET'])
def list_reservations():
    d = request.args.get('date', date.today().isoformat())
    shift = request.args.get('shift', 'comida')
    show_all = request.args.get('all', 'false') == 'true'
    target = date.fromisoformat(d)
    if show_all:
        items = res_svc.get_all_reservations_for_date(target, shift)
    else:
        items = res_svc.get_reservations(target, shift)
    return jsonify([r.to_dict() for r in items])


@api.route('/reservations/bulk-add', methods=['POST'])
@handle_errors
def import_reservation():
    """Admin-only: import reservation bypassing date validation.
    Used to enter past reservations from paper books.
    """
    data = request.json or {}
    from models import Reservation as Res
    from datetime import date as dt
    res_date = dt.fromisoformat(data['date'])
    shift = data['shift']
    time_str = data['time']
    guests = int(data['guests'])
    client_name = data['client_name'].strip()
    client_phone = data.get('client_phone', '').strip()
    notes = data.get('notes', '')
    source = data.get('source', 'phone')
    duration = int(data.get('duration_minutes', 120))

    # Find or create client
    from services import client as cli_svc
    client = None
    if client_phone:
        client = cli_svc.get_client_by_phone(client_phone)
    if not client and len(client_name) >= 2:
        try:
            client = cli_svc.create_client({'name': client_name, 'phone': client_phone})
        except Exception:
            pass

    r = Res(
        date=res_date, shift=shift, time=time_str,
        guests=guests, client_name=client_name, client_phone=client_phone,
        notes=notes, source=source, status='confirmed',
        duration_minutes=duration,
        client_id=client.id if client else None,
    )
    db.session.add(r)
    db.session.commit()
    notify()
    return jsonify(r.to_dict()), 201


@api.route('/reservations', methods=['POST'])
@validate_and_handle(ReservationCreate)
def create_reservation(data: ReservationCreate):
    """Create a new reservation.

    Request body:
        - date: YYYY-MM-DD
        - shift: 'comida' or 'cena'
        - time: HH:MM
        - guests: 1-14
        - client_name: string
        - client_phone: string
        - email: (optional) string
        - notes: (optional) string
    """
    # Convert Pydantic model to dict for service (mode='json' ensures date → ISO string)
    reservation_data = data.model_dump(mode='json')
    # Allow importing past reservations when explicitly requested
    raw = request.json or {}
    if raw.get('force_past'):
        reservation_data['force_past'] = True
    r = res_svc.create_reservation(reservation_data)
    notify(reservation_data.get('date'), reservation_data.get('shift'))
    _send_email_async(current_app._get_current_object(), reservation_data)
    return jsonify(r.to_dict()), 201


@api.route('/reservations/<int:rid>', methods=['GET'])
@handle_errors
def get_reservation(rid):
    """Get a single reservation by ID."""
    r = Reservation.query.get_or_404(rid)
    return jsonify(r.to_dict())


@api.route('/reservations/<int:rid>', methods=['PUT'])
@handle_errors
def update_reservation(rid):
    """Update an existing reservation.

    Only provided fields are updated.
    """
    if not request.is_json:
        return error_response('Content-Type must be application/json', 400)

    # Validate with partial schema (all fields optional)
    raw = request.json or {}
    try:
        data = ReservationUpdate(**{k: v for k, v in raw.items() if k != 'date'})
        update_dict = data.model_dump(mode='json', exclude_unset=True)
    except Exception as e:
        return error_response(f'Validación fallida: {str(e)}', 400)
    # Inject date directly (Pydantic v2 Optional[date] has a compat issue with @validator)
    if 'date' in raw and raw['date']:
        update_dict['date'] = raw['date']

    r = res_svc.update_reservation(rid, update_dict)
    notify(r.date.isoformat(), r.shift)
    return jsonify(r.to_dict())


@api.route('/reservations/<int:rid>', methods=['DELETE'])
@handle_errors
def cancel_reservation(rid):
    """Cancel a reservation (mark as cancelled)."""
    r = res_svc.cancel_reservation(rid)
    notify(r.date.isoformat(), r.shift)
    return jsonify(r.to_dict())


@api.route('/reservations/unassigned', methods=['GET'])
@handle_errors
def get_unassigned_reservations():
    """Future reservations without a table assigned (web bookings pending table)."""
    from app import cache_get, cache_set
    from datetime import date as _date
    cached = cache_get('unassigned', max_age=15)
    if cached:
        return jsonify(cached)
    today = _date.today()
    rows = Reservation.query.filter(
        Reservation.date >= today,
        Reservation.table_id == None,
        Reservation.status.in_(['confirmed', 'seated', 'pending']),
    ).order_by(Reservation.date, Reservation.time).all()
    result = [r.to_dict() for r in rows]
    cache_set('unassigned', result)
    return jsonify(result)


@api.route('/reservations/<int:rid>/delete', methods=['DELETE'])
@handle_errors
def delete_reservation(rid):
    """Permanently delete a reservation (hard delete)."""
    r = Reservation.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    notify()
    return jsonify({'deleted': rid})


@api.route('/reservations/<int:rid>/seat', methods=['PUT'])
@handle_errors
def seat_reservation(rid):
    """Mark reservation as seated (optionally assign table)."""
    table_id = None
    if request.is_json:
        table_id = request.json.get('table_id')

    r = res_svc.seat_reservation(rid, table_id)
    notify(r.date.isoformat(), r.shift)
    return jsonify(r.to_dict())


@api.route('/reservations/<int:rid>/complete', methods=['PUT'])
@handle_errors
def complete_reservation(rid):
    """Mark reservation as completed."""
    r = res_svc.complete_reservation(rid)
    notify(r.date.isoformat(), r.shift)
    return jsonify(r.to_dict())


@api.route('/reservations/<int:rid>/noshow', methods=['PUT'])
@handle_errors
def mark_no_show(rid):
    """Mark reservation as no-show."""
    r = res_svc.mark_no_show(rid)
    notify(r.date.isoformat(), r.shift)
    return jsonify(r.to_dict())


@api.route('/reservations/bulk-complete', methods=['POST'])
@handle_errors
def bulk_complete_reservations():
    """Complete (free) several SEATED reservations at once.

    Body:
      - {ids: [1,2,3]}                 → complete exactly those (only if seated)
      - {date, shift}  (no ids)        → complete ALL seated of that date+shift
    Only touches status == 'seated' (mesas ocupadas), never confirmadas/futuras.
    Returns {completed: N}.
    """
    data = request.json or {}
    ids = data.get('ids')

    q = Reservation.query.filter(Reservation.status == 'seated')
    if ids:
        try:
            id_list = [int(x) for x in ids if x]
        except (TypeError, ValueError):
            return error_response('ids inválidos', 400)
        if not id_list:
            return jsonify({'completed': 0})
        q = q.filter(Reservation.id.in_(id_list))
    else:
        d = data.get('date')
        shift = data.get('shift')
        if d:
            q = q.filter(Reservation.date == date.fromisoformat(d))
        if shift:
            q = q.filter(Reservation.shift == shift)

    rows = q.all()
    for r in rows:
        r.status = 'completed'
    if rows:
        db.session.commit()
        notify()
    return jsonify({'completed': len(rows)})


# ── Tables ────────────────────────────────────────────────

@api.route('/tables', methods=['GET'])
def list_tables():
    tables = Table.query.filter_by(active=True).order_by(Table.zone, Table.number).all()
    return jsonify([t.to_dict() for t in tables])


@api.route('/tables/status', methods=['GET'])
def table_status():
    from app import cache_get, cache_set
    d = request.args.get('date', date.today().isoformat())
    shift = request.args.get('shift', 'comida')
    cache_key = f'tables_status:{d}:{shift}'
    cached = cache_get(cache_key)
    if cached:
        return jsonify(cached)
    target = date.fromisoformat(d)
    result = res_svc.get_table_status(target, shift)
    cache_set(cache_key, result)
    return jsonify(result)


@api.route('/tables', methods=['POST'])
@handle_errors
def create_table():
    """Create a new table. Body: { number, capacity, zone, table_type?, pos_x?, pos_y?, width?, height? }"""
    data = request.json or {}
    number = data.get('number')
    if not number:
        return error_response('Número de mesa requerido', 400)
    number = int(number)
    if Table.query.filter_by(number=number).first():
        return error_response(f'La mesa {number} ya existe', 400)
    capacity = int(data.get('capacity', 4))
    if capacity < 1 or capacity > 30:
        return error_response('Capacidad debe estar entre 1 y 30', 400)
    zone = data.get('zone', 'sp')
    table_type = data.get('table_type', 'normal')
    pos_x = float(data.get('pos_x', 50))
    pos_y = float(data.get('pos_y', 50))
    shape = data.get('shape', 'rect')
    if shape not in ('rect', 'square', 'circle'):
        shape = 'rect'
    t = Table(
        number=number, capacity=capacity, zone=zone,
        table_type=table_type, pos_x=pos_x, pos_y=pos_y,
        shape=shape, sofa_type='', sofa_side='',
        active=True, blocked=False,
    )
    db.session.add(t)
    db.session.commit()
    notify()
    return jsonify(t.to_dict()), 201


@api.route('/tables/<int:tid>', methods=['DELETE'])
@handle_errors
def delete_table(tid):
    """Soft-delete a table (active=False). Blocked if it has upcoming active reservations."""
    t = Table.query.get_or_404(tid)
    future_res = Reservation.query.filter(
        Reservation.table_id == tid,
        Reservation.date >= date.today(),
        Reservation.status.in_(['confirmed', 'seated'])
    ).count()
    if future_res > 0:
        return error_response(f'La mesa tiene {future_res} reserva(s) activa(s) — cancélalas primero', 400)
    t.active = False
    db.session.commit()
    notify()
    return jsonify({'ok': True, 'id': tid})


@api.route('/tables/<int:tid>', methods=['PUT'])
@handle_errors
def update_table(tid):
    """Update table capacity, type and/or position.
    Body: { capacity?: int, table_type?: str, pos_x?: float, pos_y?: float }
    """
    t = Table.query.get_or_404(tid)
    data = request.json or {}
    if 'capacity' in data:
        cap = int(data['capacity'])
        if cap < 1 or cap > 30:
            return error_response('Capacidad debe estar entre 1 y 30', 400)
        t.capacity = cap
    if 'table_type' in data and data['table_type'] in ('normal', 'alta'):
        t.table_type = data['table_type']
    if 'pos_x' in data:
        t.pos_x = float(data['pos_x'])
    if 'pos_y' in data:
        t.pos_y = float(data['pos_y'])
    if 'svg_w' in data:
        t.svg_w = float(data['svg_w']) if data['svg_w'] is not None else None
    if 'svg_h' in data:
        t.svg_h = float(data['svg_h']) if data['svg_h'] is not None else None
    if 'shape' in data and data['shape'] in ('rect', 'square', 'circle'):
        t.shape = data['shape']
    if 'sofa_type' in data and data['sofa_type'] in ('', 'straight', 'L'):
        t.sofa_type = data['sofa_type']
    if 'sofa_side' in data:
        t.sofa_side = (data['sofa_side'] or '')[:10]
    if 'blocked' in data:
        t.blocked = bool(data['blocked'])
    db.session.commit()
    notify()
    return jsonify(t.to_dict())


@api.route('/tables/available', methods=['GET'])
def available_tables():
    d = request.args.get('date', date.today().isoformat())
    shift = request.args.get('shift', 'comida')
    guests = int(request.args.get('guests', 2))
    time_str = request.args.get('time')
    duration = int(request.args.get('duration', 120))
    exclude_id = request.args.get('exclude_reservation_id', type=int)
    target = date.fromisoformat(d)
    tables = res_svc.find_available_tables(target, shift, guests, time_str, duration, exclude_id)
    return jsonify([t.to_dict() for t in tables])


@api.route('/tables/<int:tid>/quick-occupy', methods=['POST'])
@handle_errors
def quick_occupy(tid):
    """Mark a table as occupied instantly — no client name needed.
    Body: { guests: int, shift: str }
    """
    table = Table.query.get_or_404(tid)
    if table.blocked:
        return error_response('Mesa congelada — desbloquéala primero', 400)
    data = request.json or {}
    guests = max(1, int(data.get('guests', 2)))
    shift = data.get('shift', 'comida')
    now_str = datetime.now().strftime('%H:%M')
    today = date.today()

    # If a quick-occupy reservation already exists for this table today, update it
    existing = Reservation.query.filter_by(
        table_id=tid,
        date=today,
        shift=shift,
        source='walk_in',
    ).filter(Reservation.status.in_(['confirmed', 'seated'])).first()

    if existing:
        existing.guests = guests
        if existing.status != 'seated' or not existing.seated_at:
            existing.seated_at = datetime.utcnow()
        existing.status = 'seated'
        db.session.commit()
        notify()
        return jsonify({'ok': True, 'reservation_id': existing.id})

    r = Reservation(
        date=today,
        shift=shift,
        time=now_str,
        guests=guests,
        client_name='Walk-in',
        client_phone='',
        table_id=table.id,
        status='seated',
        source='walk_in',
        notes='Ocupación directa desde plano',
        duration_minutes=120,
        seated_at=datetime.utcnow(),
    )
    db.session.add(r)
    db.session.commit()
    notify()
    return jsonify({'ok': True, 'reservation_id': r.id}), 201


@api.route('/tables/<int:tid>/quick-free', methods=['PUT'])
@handle_errors
def quick_free(tid):
    """Free a table — marks the active SEATED reservation as completed.

    Prefers a seated sitting over a merely confirmed (future) one so that on a
    table with two sittings we never complete the wrong reservation. The floor
    plan frees walk-ins by reservation id (/reservations/<id>/complete); this
    endpoint is kept as a defensive fallback.
    """
    data = request.json or {}
    shift = data.get('shift', 'comida')
    today = date.today()

    r = Reservation.query.filter_by(
        table_id=tid,
        date=today,
        shift=shift,
    ).filter(
        Reservation.status.in_(['confirmed', 'seated'])
    ).order_by(
        # 'seated' > 'confirmed' alphabetically → desc() puts seated first
        Reservation.status.desc()
    ).first()

    if r:
        r.status = 'completed'
        db.session.commit()
        notify()

    return jsonify({'ok': True})


@api.route('/tables/check-conflict', methods=['POST'])
@handle_errors
def check_table_conflict():
    """Check if a set of tables is available at a given time. Used by frontend before submit.

    Body: {date, shift, time, duration_minutes, table_ids, exclude_reservation_id}
    Returns: {conflicts: [...], available: bool}
    """
    data = request.json or {}
    target = date.fromisoformat(data.get('date'))
    conflicts = res_svc.check_table_conflicts(
        target,
        data.get('shift'),
        data.get('time'),
        int(data.get('duration_minutes', 120) or 120),
        data.get('table_ids', []),
        data.get('exclude_reservation_id'),
    )
    return jsonify({
        'available': len(conflicts) == 0,
        'conflicts': conflicts,
    })


# ── Stats ─────────────────────────────────────────────────

@api.route('/stats', methods=['GET'])
def shift_stats():
    d = request.args.get('date', date.today().isoformat())
    shift = request.args.get('shift', 'comida')
    target = date.fromisoformat(d)
    stats = res_svc.get_shift_stats(target, shift)
    return jsonify(stats)


@api.route('/quick-status', methods=['GET'])
def quick_status():
    """Ultra-light: only stats, cached 3s. For mobile quick updates."""
    from app import cache_get, cache_set
    d = request.args.get('date', date.today().isoformat())
    shift = request.args.get('shift', 'comida')
    cache_key = f'stats:{d}:{shift}'

    cached = cache_get(cache_key, max_age=3)
    if cached:
        return jsonify(cached)

    target = date.fromisoformat(d)
    stats = res_svc.get_shift_stats(target, shift)
    cache_set(cache_key, stats)
    return jsonify(stats)


@api.route('/dashboard', methods=['GET'])
def dashboard():
    """Stats + table status + reservations in one call. 2 queries total (was 4)."""
    from app import cache_get, cache_set
    d = request.args.get('date', date.today().isoformat())
    shift = request.args.get('shift', 'comida')
    cache_key = f'dashboard:{d}:{shift}'

    cached = cache_get(cache_key, max_age=30)
    if cached:
        return jsonify(cached)

    target = date.fromisoformat(d)

    # Query 1: all reservations for the shift (eager-loaded, no N+1)
    all_res = res_svc.get_all_reservations_for_date(target, shift)
    reservations = [r.to_dict() for r in all_res]

    # Derive stats from already-loaded data — no extra query
    stats = res_svc.stats_from_reservations(reservations)

    # Query 2: table status (reuses same reservations internally via eager load)
    tables = res_svc.get_table_status(target, shift)
    slim_tables = [
        {'id': t['id'], 'number': t['number'], 'status': t['status'],
         'blocked': t.get('blocked', False), 'reservation': t.get('reservation'),
         'shape': t.get('shape', 'rect'),
         'sofa_type': t.get('sofa_type', ''), 'sofa_side': t.get('sofa_side', ''),
         'table_type': t.get('table_type', 'normal'),
         'svg_w': t.get('svg_w'), 'svg_h': t.get('svg_h')}
        for t in tables if 'id' in t
    ]

    result = {'stats': stats, 'tables': slim_tables, 'reservations': reservations}
    cache_set(cache_key, result)
    return jsonify(result)


# ── Clients ───────────────────────────────────────────────

@api.route('/clients', methods=['GET'])
@handle_errors
def list_clients():
    """Search clients by name, phone, or email."""
    q = request.args.get('search', '')
    clients = cli_svc.search_clients(q)
    return jsonify([c.to_dict() for c in clients])


@api.route('/clients/<int:cid>', methods=['GET'])
@handle_errors
def get_client(cid):
    """Get client details with reservation history."""
    c = cli_svc.get_client(cid)
    history = cli_svc.get_client_history(cid)
    return jsonify({
        **c.to_dict(),
        'history': [r.to_dict() for r in history],
    })


@api.route('/clients', methods=['POST'])
@validate_and_handle(ClientCreate)
def create_client(data: ClientCreate):
    """Create a new client."""
    c = cli_svc.create_client(data.dict())
    return jsonify(c.to_dict()), 201


@api.route('/clients/<int:cid>', methods=['PUT'])
@handle_errors
def update_client(cid):
    """Update an existing client."""
    if not request.is_json:
        return error_response('Content-Type must be application/json', 400)

    try:
        data = ClientUpdate(**request.json)
        update_dict = data.dict(exclude_unset=True)
    except Exception as e:
        return error_response(f'Validación fallida: {str(e)}', 400)

    c = cli_svc.update_client(cid, update_dict)
    return jsonify(c.to_dict())


@api.route('/clients/<int:cid>', methods=['DELETE'])
@handle_errors
def delete_client(cid):
    """Delete a client (detach reservations)."""
    cli_svc.delete_client(cid)
    return jsonify({'status': 'deleted'})


@api.route('/clients/top', methods=['GET'])
def top_clients():
    limit = int(request.args.get('limit', 20))
    return jsonify(cli_svc.get_top_clients(limit))


# ── Waitlist ──────────────────────────────────────────────

@api.route('/waitlist', methods=['GET'])
def list_waitlist():
    d = request.args.get('date', date.today().isoformat())
    shift = request.args.get('shift', 'comida')
    items = Waitlist.query.filter(
        Waitlist.date == date.fromisoformat(d),
        Waitlist.shift == shift,
        Waitlist.status == 'waiting',
    ).order_by(Waitlist.created_at).all()
    return jsonify([w.to_dict() for w in items])


@api.route('/waitlist', methods=['POST'])
def add_to_waitlist():
    data = request.json
    w = Waitlist(
        client_name=data['client_name'],
        phone=data.get('phone', ''),
        guests=int(data['guests']),
        date=date.fromisoformat(data['date']),
        shift=data['shift'],
        notes=data.get('notes', ''),
    )
    db.session.add(w)
    db.session.commit()
    notify()
    return jsonify(w.to_dict()), 201


@api.route('/waitlist/<int:wid>/seat', methods=['PUT'])
def seat_from_waitlist(wid):
    w = Waitlist.query.get_or_404(wid)
    data = request.json or {}
    r = res_svc.create_reservation({
        'client_name': w.client_name,
        'client_phone': w.phone,
        'date': w.date.isoformat(),
        'shift': w.shift,
        'time': data.get('time', ''),
        'guests': w.guests,
        'table_id': data.get('table_id'),
        'source': 'walk_in',
        'notes': w.notes,
    })
    w.status = 'seated'
    db.session.commit()
    notify()
    return jsonify(r.to_dict()), 201


@api.route('/waitlist/<int:wid>/cancel', methods=['PUT'])
def cancel_waitlist(wid):
    w = Waitlist.query.get_or_404(wid)
    w.status = 'cancelled'
    db.session.commit()
    notify()
    return jsonify(w.to_dict())


# ── Config / WhatsApp ─────────────────────────────────────

@api.route('/config/whatsapp', methods=['GET'])
def whatsapp_config():
    return jsonify({
        'number': current_app.config['WHATSAPP_NUMBER'],
        'link': f"https://wa.me/{current_app.config['WHATSAPP_NUMBER']}?text=RESERVAR",
    })


# ── Reports ───────────────────────────────────────────────

@api.route('/reports/occupancy', methods=['GET'])
@cached_report()
def report_occupancy():
    from_d = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    return jsonify(rpt_svc.occupancy_report(date.fromisoformat(from_d), date.fromisoformat(to_d)))


@api.route('/reports/popular-tables', methods=['GET'])
@cached_report()
def report_popular_tables():
    from_d = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    return jsonify(rpt_svc.popular_tables_report(date.fromisoformat(from_d), date.fromisoformat(to_d)))


@api.route('/reports/sources', methods=['GET'])
@cached_report()
def report_sources():
    from_d = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    return jsonify(rpt_svc.source_report(date.fromisoformat(from_d), date.fromisoformat(to_d)))


@api.route('/reports/no-shows', methods=['GET'])
@cached_report()
def report_no_shows():
    from_d = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    return jsonify(rpt_svc.no_show_report(date.fromisoformat(from_d), date.fromisoformat(to_d)))


@api.route('/reports/summary', methods=['GET'])
@cached_report(max_age=30)
def report_summary():
    d = request.args.get('date', date.today().isoformat())
    return jsonify(rpt_svc.daily_summary(date.fromisoformat(d)))


@api.route('/reports/weekly', methods=['GET'])
@cached_report()
def report_weekly():
    from_d = request.args.get('from', date.today().isoformat())
    return jsonify(rpt_svc.weekly_trend(date.fromisoformat(from_d)))


@api.route('/reports/range', methods=['GET'])
def report_range():
    """Return daily summary for a date range (used by agenda month view)."""
    from app import cache_get, cache_set
    from_d = request.args.get('from', date.today().isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    cache_key = f'reports_range:{from_d}:{to_d}'
    cached = cache_get(cache_key, max_age=30)
    if cached:
        return jsonify(cached)
    start = date.fromisoformat(from_d)
    end = date.fromisoformat(to_d)
    days = (end - start).days + 1
    if days > 42:
        days = 42
    result = rpt_svc.weekly_trend(start, days=days)
    cache_set(cache_key, result)
    return jsonify(result)


@api.route('/reports/new-vs-returning', methods=['GET'])
@cached_report()
def report_new_vs_returning():
    from_d = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    return jsonify(rpt_svc.new_vs_returning(date.fromisoformat(from_d), date.fromisoformat(to_d)))


@api.route('/reports/week-comparison', methods=['GET'])
@cached_report()
def report_week_comparison():
    # Monday of the requested week (default: current week)
    from_d = request.args.get('from', date.today().isoformat())
    target = date.fromisoformat(from_d)
    # Find Monday of that week
    monday = target - timedelta(days=target.weekday())
    return jsonify(rpt_svc.weekly_comparison(monday))


@api.route('/reports/clients', methods=['GET'])
@cached_report()
def report_clients():
    from_d = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    limit = int(request.args.get('limit', 30))
    return jsonify(rpt_svc.client_frequency_report(date.fromisoformat(from_d), date.fromisoformat(to_d), limit))


@api.route('/reports/zones', methods=['GET'])
@cached_report()
def report_zones():
    from_d = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    return jsonify(rpt_svc.zone_performance(date.fromisoformat(from_d), date.fromisoformat(to_d)))


@api.route('/reports/hours', methods=['GET'])
@cached_report()
def report_hours():
    from_d = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    return jsonify(rpt_svc.hourly_distribution(date.fromisoformat(from_d), date.fromisoformat(to_d)))


@api.route('/reports/kpi', methods=['GET'])
@cached_report()
def report_kpi():
    from_d = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    return jsonify(rpt_svc.kpi_summary(date.fromisoformat(from_d), date.fromisoformat(to_d)))


@api.route('/reports/heatmap', methods=['GET'])
@cached_report()
def report_heatmap():
    from_d = request.args.get('from', (date.today() - timedelta(days=90)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    return jsonify(rpt_svc.heatmap_report(date.fromisoformat(from_d), date.fromisoformat(to_d)))


@api.route('/reports/trend', methods=['GET'])
@cached_report()
def report_trend():
    from_d = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    return jsonify(rpt_svc.weekly_occupancy(date.fromisoformat(from_d), date.fromisoformat(to_d)))


@api.route('/reservations/day', methods=['GET'])
def list_reservations_day():
    """Return all reservations for a date (both shifts), sorted by time."""
    d = request.args.get('date', date.today().isoformat())
    target = date.fromisoformat(d)
    from models import Reservation as Res
    comida = Res.query.filter_by(date=target, shift='comida').order_by(Res.time).all()
    cena = Res.query.filter_by(date=target, shift='cena').order_by(Res.time).all()
    return jsonify({
        'comida': [r.to_dict() for r in comida],
        'cena': [r.to_dict() for r in cena],
        'date': d,
        'stats': {
            'comida_total': len(comida),
            'comida_guests': sum(r.guests for r in comida),
            'comida_active': sum(1 for r in comida if r.status in ('confirmed', 'seated', 'completed')),
            'cena_total': len(cena),
            'cena_guests': sum(r.guests for r in cena),
            'cena_active': sum(1 for r in cena if r.status in ('confirmed', 'seated', 'completed')),
        }
    })


# ═══════════════════════════════════════════════════════════════════════════
# GLOBAL SEARCH (across all dates)
# ═══════════════════════════════════════════════════════════════════════════

@api.route('/search/reservations', methods=['GET'])
@handle_errors
def search_reservations():
    """Search reservations across all dates by name/phone/notes."""
    q = request.args.get('q', '').strip()
    limit = int(request.args.get('limit', 50))
    if len(q) < 2:
        return jsonify([])
    results = res_svc.global_search_reservations(q, limit)
    return jsonify([r.to_dict() for r in results])


# ═══════════════════════════════════════════════════════════════════════════
# EXPORT CSV
# ═══════════════════════════════════════════════════════════════════════════

@api.route('/export/reservations.csv', methods=['GET'])
def export_reservations_csv():
    """Export reservations between two dates as CSV (Excel-compatible)."""
    from_d = request.args.get('from', date.today().isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    from_date = date.fromisoformat(from_d)
    to_date = date.fromisoformat(to_d)

    items = Reservation.query.filter(
        Reservation.date >= from_date,
        Reservation.date <= to_date,
    ).order_by(Reservation.date, Reservation.time).all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow([
        'ID', 'Fecha', 'Turno', 'Hora', 'Cliente', 'Telefono',
        'Comensales', 'Mesas', 'Estado', 'Origen', 'Duracion(min)', 'Notas', 'Creada'
    ])
    for r in items:
        nums = r.all_table_numbers()
        tables_str = ' + '.join(str(n) for n in nums) if nums else ''
        writer.writerow([
            r.id, r.date.isoformat(), r.shift, r.time,
            r.client_name, r.client_phone, r.guests, tables_str,
            r.status, r.source, r.duration_minutes or 120,
            (r.notes or '').replace('\n', ' '),
            r.created_at.isoformat() if r.created_at else ''
        ])

    csv_data = '﻿' + output.getvalue()  # BOM for Excel UTF-8
    output.close()

    return Response(
        csv_data,
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f'attachment; filename=reservas_{from_d}_a_{to_d}.csv'
        }
    )


@api.route('/export/backup.xlsx', methods=['GET'])
def export_backup_xlsx():
    """Copia de seguridad: TODAS las reservas de la base de datos en un Excel
    de una sola hoja (histórico completo). Un clic, sin rango de fechas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    status_map = {'confirmed': 'Confirmada', 'seated': 'Sentada', 'completed': 'Completada',
                  'cancelled': 'Cancelada', 'no_show': 'No Show', 'pending': 'Pendiente'}
    source_map = {'phone': 'Teléfono', 'whatsapp': 'WhatsApp', 'walk_in': 'Walk-in', 'web': 'Web'}

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reservas'
    ws.sheet_view.showGridLines = False

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A8A7D')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    today_str = date.today().isoformat()
    ws['A1'].value = f'COPIA DE SEGURIDAD — TODAS LAS RESERVAS · generado {today_str}'
    ws['A1'].font = Font(bold=True, size=13, color='111827')
    ws.merge_cells('A1:M1')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    headers = ['ID', 'Fecha', 'Turno', 'Hora', 'Cliente', 'Teléfono',
               'Comensales', 'Mesa(s)', 'Estado', 'Origen', 'Duración', 'Notas', 'Creada']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = 'A3'  # cabecera fija al hacer scroll

    items = Reservation.query.order_by(Reservation.date, Reservation.time).all()
    for idx, r in enumerate(items, 3):
        nums = r.all_table_numbers()
        row_data = [
            r.id, r.date.isoformat(), (r.shift or '').capitalize(), r.time,
            r.client_name, r.client_phone or '',
            r.guests, ' + '.join(str(n) for n in nums) if nums else 'Sin mesa',
            status_map.get(r.status, r.status), source_map.get(r.source, r.source or ''),
            f"{r.duration_minutes or 120} min", r.notes or '',
            r.created_at.strftime('%d/%m/%Y %H:%M') if r.created_at else ''
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=j, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F9FAFB')

    for i, w in enumerate([6, 12, 10, 8, 22, 14, 11, 12, 12, 12, 10, 24, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=copia_reservas_season_{today_str}.xlsx'}
    )


@api.route('/export/report.xlsx', methods=['GET'])
def export_report_xlsx():
    """Export a full analytics report as Excel (.xlsx) with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    from_d = request.args.get('from', (date.today() - timedelta(days=30)).isoformat())
    to_d = request.args.get('to', date.today().isoformat())
    from_date_d = date.fromisoformat(from_d)
    to_date_d = date.fromisoformat(to_d)

    wb = Workbook()

    # ── Styles ──────────────────────────────────
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A8A7D')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    def style_data_row(ws, row, cols, alt=False):
        fill = PatternFill('solid', fgColor='F9FAFB') if alt else None
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if fill: cell.fill = fill

    # ── HOJA 1: Reservas detalladas ──────────────
    ws1 = wb.active
    ws1.title = 'Reservas'
    ws1.sheet_view.showGridLines = False

    title_cell = ws1['A1']
    title_cell.value = f'RESERVAS — {from_d} a {to_d}'
    title_cell.font = Font(bold=True, size=13, color='111827')
    ws1.merge_cells('A1:M1')
    ws1['A1'].alignment = center
    ws1.row_dimensions[1].height = 28

    headers1 = ['ID', 'Fecha', 'Turno', 'Hora', 'Cliente', 'Teléfono',
                'Comensales', 'Mesa(s)', 'Estado', 'Origen', 'Duración', 'Notas', 'Creada']
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=2, column=i, value=h)
    style_header_row(ws1, 2, len(headers1))
    ws1.row_dimensions[2].height = 20

    status_map = {'confirmed': 'Confirmada', 'seated': 'Sentada', 'completed': 'Completada',
                  'cancelled': 'Cancelada', 'no_show': 'No Show'}
    source_map = {'phone': 'Teléfono', 'whatsapp': 'WhatsApp', 'walk_in': 'Walk-in', 'web': 'Web'}

    items = Reservation.query.filter(
        Reservation.date >= from_date_d, Reservation.date <= to_date_d
    ).order_by(Reservation.date, Reservation.time).all()

    for i, r in enumerate(items, 3):
        nums = r.all_table_numbers()
        row_data = [
            r.id, r.date.isoformat(), r.shift.capitalize(), r.time,
            r.client_name, r.client_phone or '',
            r.guests, ' + '.join(str(n) for n in nums) if nums else 'Sin mesa',
            status_map.get(r.status, r.status), source_map.get(r.source, r.source or ''),
            f"{r.duration_minutes or 120} min", r.notes or '',
            r.created_at.strftime('%d/%m/%Y %H:%M') if r.created_at else ''
        ]
        for j, val in enumerate(row_data, 1):
            ws1.cell(row=i, column=j, value=val)
        style_data_row(ws1, i, len(headers1), alt=(i % 2 == 0))

    col_widths1 = [6, 12, 10, 8, 22, 14, 11, 12, 12, 12, 10, 24, 16]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── HOJA 2: Resumen por día ──────────────────
    ws2 = wb.create_sheet('Resumen por día')
    ws2.sheet_view.showGridLines = False
    ws2['A1'].value = f'RESUMEN POR DÍA — {from_d} a {to_d}'
    ws2['A1'].font = Font(bold=True, size=13, color='111827')
    ws2.merge_cells('A1:H1')
    ws2['A1'].alignment = center
    ws2.row_dimensions[1].height = 28

    headers2 = ['Fecha', 'Día', 'Reservas', 'Comensales', 'Comida (p)', 'Cena (p)', 'Ocupación %', 'No Shows']
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header_row(ws2, 2, len(headers2))

    trend = rpt_svc.weekly_occupancy(from_date_d, to_date_d)
    days_es = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    for i, d_data in enumerate(trend, 3):
        dt = date.fromisoformat(d_data['date'])
        noshows = Reservation.query.filter_by(date=dt, status='no_show').count()
        row_data = [
            dt.strftime('%d/%m/%Y'),
            days_es[dt.weekday()],
            d_data['reservations'], d_data['guests'],
            d_data.get('comida', 0), d_data.get('cena', 0),
            f"{d_data['occupancy']}%", noshows
        ]
        for j, val in enumerate(row_data, 1):
            ws2.cell(row=i, column=j, value=val)
        style_data_row(ws2, i, len(headers2), alt=(i % 2 == 0))

    for i, w in enumerate([14, 12, 10, 12, 12, 10, 12, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── HOJA 3: KPIs del período ─────────────────
    ws3 = wb.create_sheet('KPIs')
    ws3.sheet_view.showGridLines = False
    ws3['A1'].value = f'MÉTRICAS — {from_d} a {to_d}'
    ws3['A1'].font = Font(bold=True, size=13, color='111827')
    ws3.merge_cells('A1:C1')
    ws3['A1'].alignment = center
    ws3.row_dimensions[1].height = 28

    kpi = rpt_svc.kpi_summary(from_date_d, to_date_d)
    curr = kpi['current']
    kpi_rows = [
        ('Métrica', 'Período actual', 'Período anterior'),
        ('Total reservas', curr['total'], kpi['previous']['total']),
        ('Total comensales', curr['guests'], kpi['previous']['guests']),
        ('Ocupación media (%)', curr['occupancy_avg'], kpi['previous']['occupancy_avg']),
        ('Tasa No-Show (%)', curr['no_show_rate'], kpi['previous']['no_show_rate']),
        ('Media comensales/reserva', curr['avg_guests'], kpi['previous']['avg_guests']),
        ('Reservas completadas', curr['completed'], kpi['previous']['completed']),
        ('Reservas canceladas', curr['cancelled'], kpi['previous']['cancelled']),
    ]
    for i, row in enumerate(kpi_rows, 2):
        for j, val in enumerate(row, 1):
            ws3.cell(row=i, column=j, value=val)
        if i == 2:
            style_header_row(ws3, i, 3)
        else:
            style_data_row(ws3, i, 3, alt=(i % 2 == 0))
    for i, w in enumerate([28, 18, 18], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── HOJA 4: Top Clientes ─────────────────────
    ws4 = wb.create_sheet('Top Clientes')
    ws4.sheet_view.showGridLines = False
    ws4['A1'].value = f'TOP CLIENTES — {from_d} a {to_d}'
    ws4['A1'].font = Font(bold=True, size=13, color='111827')
    ws4.merge_cells('A1:F1')
    ws4['A1'].alignment = center
    ws4.row_dimensions[1].height = 28

    headers4 = ['#', 'Cliente', 'Teléfono', 'Visitas', 'Total comensales', 'Última visita']
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=2, column=i, value=h)
    style_header_row(ws4, 2, len(headers4))

    clients = rpt_svc.client_frequency_report(from_date_d, to_date_d, limit=30)
    for i, c in enumerate(clients, 3):
        row_data = [i-2, c['name'], c['phone'] or '', c['visits'], c['total_guests'],
                    c['last_visit'] if c['last_visit'] else '']
        for j, val in enumerate(row_data, 1):
            ws4.cell(row=i, column=j, value=val)
        style_data_row(ws4, i, len(headers4), alt=(i % 2 == 0))
    for i, w in enumerate([5, 24, 16, 10, 16, 14], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ── Return file ──────────────────────────────
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=informe_season_{from_d}_a_{to_d}.xlsx'}
    )


# ═══════════════════════════════════════════════════════════════════════════
# DRAG-DROP: ASSIGN RESERVATION TO TABLE(S)
# ═══════════════════════════════════════════════════════════════════════════

@api.route('/reservations/<int:rid>/assign-tables', methods=['PUT'])
@handle_errors
def assign_tables(rid):
    """Directly assign table(s) to a reservation (floor plan drag-drop / move).

    Bypasses conflict and capacity checks — used by staff from the floor plan
    where free-table filtering is already handled in the UI.
    Body: {table_ids: [1,2]} or {table_id: 1}
    """
    data = request.json or {}
    reservation = Reservation.query.get_or_404(rid)

    if data.get('table_id'):
        new_table_id = int(data['table_id'])
        new_table = Table.query.get_or_404(new_table_id)
        reservation.table_id = new_table_id
        reservation.extra_tables = [new_table]
    elif 'table_ids' in data:
        table_ids = [int(x) for x in (data['table_ids'] or []) if x]
        if table_ids:
            tables = Table.query.filter(Table.id.in_(table_ids)).all()
            reservation.table_id = table_ids[0]
            reservation.extra_tables = tables
        else:
            reservation.table_id = None
            reservation.extra_tables = []

    db.session.commit()
    notify()
    return jsonify(reservation.to_dict())


@api.route('/reservations/<int:rid>/unassign', methods=['PUT'])
@handle_errors
def unassign_tables(rid):
    """Remove all table assignments from a reservation."""
    r = res_svc.update_reservation(rid, {'table_ids': []})
    notify()
    return jsonify(r.to_dict())


# ═══════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════
# APP SETTINGS
# ═══════════════════════════════════════════════════════════════════════════

@api.route('/settings/<key>', methods=['GET'])
@handle_errors
def get_setting(key):
    from models import AppSetting
    ALLOWED = {'web_capacity_limit'}
    if key not in ALLOWED:
        return error_response('Setting no permitido', 403)
    value = AppSetting.get(key, None)
    return jsonify({'key': key, 'value': value})

@api.route('/settings/<key>', methods=['PUT'])
@handle_errors
def put_setting(key):
    from models import AppSetting
    ALLOWED = {'web_capacity_limit'}
    if key not in ALLOWED:
        return error_response('Setting no permitido', 403)
    data = request.json or {}
    value = data.get('value')
    if value is None:
        return error_response('Falta value', 400)
    if key == 'web_capacity_limit':
        try:
            v = int(value)
            if v < 1 or v > 500:
                raise ValueError
        except (ValueError, TypeError):
            return error_response('El límite debe ser un número entre 1 y 500', 400)
    AppSetting.set(key, value)
    return jsonify({'key': key, 'value': str(value)})


# PUBLIC BOOKING (no auth required — used by /reservas page)
# ═══════════════════════════════════════════════════════════════════════════

@api.route('/public/reserve', methods=['POST'])
@handle_errors
def public_reserve():
    """Create a reservation from the public booking page. No auth required."""
    data = request.json or {}

    required = ['client_name', 'client_phone', 'date', 'shift', 'time', 'guests']
    missing = [f for f in required if not data.get(f)]
    if missing:
        return error_response(f'Faltan campos: {", ".join(missing)}', 400)

    if data['shift'] not in ('comida', 'cena'):
        return error_response('Turno debe ser comida o cena', 400)

    try:
        guests = int(data['guests'])
        if guests < 1 or guests > 20:
            raise ValueError
    except (ValueError, TypeError):
        return error_response('Número de comensales inválido (1-8)', 400)

    if guests > 8:
        return error_response(
            'Grupos de más de 8 personas: contacta directamente con nosotros antes de reservar '
            'para que podamos preparar el espacio adecuado. Tel: +34 689 135 630', 400)

    # Overbooking check: limit comes from DB setting (default 30)
    from models import AppSetting as _AS
    try:
        MAX_SHIFT_GUESTS = int(_AS.get('web_capacity_limit', 30))
    except (ValueError, TypeError):
        MAX_SHIFT_GUESTS = 30
    from models import Reservation
    from sqlalchemy import func as _func
    from datetime import date as _date
    try:
        booked = db.session.query(_func.sum(Reservation.guests)).filter(
            Reservation.date == _date.fromisoformat(data['date']),
            Reservation.shift == data['shift'],
            Reservation.status.in_(['confirmed', 'seated']),
        ).scalar() or 0
        if booked + guests > MAX_SHIFT_GUESTS:
            remaining = max(0, MAX_SHIFT_GUESTS - booked)
            if remaining == 0:
                return error_response(
                    f'Lo sentimos, el turno de {data["shift"]} de ese día está completo. '
                    f'Por favor elige otro día o contáctanos por WhatsApp o llámanos al +34 689 135 630.', 409)
            return error_response(
                f'Solo quedan {remaining} plazas disponibles en ese turno. '
                f'Reduce el número de comensales o contáctanos por WhatsApp o llámanos al +34 689 135 630.', 409)
    except Exception:
        pass  # If check fails, allow the booking rather than blocking it

    # Find or create client
    client = cli_svc.get_client_by_phone(data['client_phone'].strip())
    if not client:
        try:
            client = cli_svc.create_client({
                'name': data['client_name'].strip(),
                'phone': data['client_phone'].strip(),
            })
        except Exception:
            client = None

    pub_data = {
        'date': data['date'],
        'shift': data['shift'],
        'time': data['time'],
        'guests': guests,
        'client_name': data['client_name'].strip(),
        'client_phone': data['client_phone'].strip(),
        'email': data.get('email', ''),
        'notes': data.get('notes', ''),
        'source': 'web',
        'client_id': client.id if client else None,
    }
    r = res_svc.create_reservation(pub_data)
    notify()
    _send_email_async(current_app._get_current_object(), pub_data)
    return jsonify(r.to_dict()), 201
